import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Request, Query
from fastapi.responses import StreamingResponse

from auth import require_auth
from db import list_gastos, list_categorias, monthly_summary, detect_transfers, mark_transfers, update_categoria, update_usuario, update_gasto_fecha, delete_all_gastos, get_gasto, delete_gasto_manual, list_importaciones, rename_categoria_in_gastos

router = APIRouter()


def _parse_categorias(categorias: Optional[str]) -> Optional[list]:
    if not categorias:
        return None
    parts = [c.strip() for c in categorias.split(",") if c.strip()]
    return parts or None


@router.get("/gastos/monthly")
def get_monthly(request: Request):
    require_auth(request)
    return monthly_summary()


@router.get("/importaciones")
def get_importaciones(request: Request):
    require_auth(request)
    return list_importaciones()


@router.get("/categorias")
def get_categorias(request: Request):
    require_auth(request)
    return list_categorias()


@router.post("/categorias/rename")
def post_rename_categoria(body: dict, request: Request):
    """Rename or clear a category across all gastos.
    {old: "OldName", new: "NewName"} — pass new="" to clear."""
    require_auth(request)
    old = str(body.get("old", "")).strip()
    new = str(body.get("new", "")).strip()
    if not old:
        return {"actualizados": 0}
    count = rename_categoria_in_gastos(old, new or None)
    return {"actualizados": count}


@router.get("/gastos/export")
def export_gastos(
    request: Request,
    fuente: Optional[str] = Query(None),
    categorias: Optional[str] = Query(None),
    usuario: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    moneda: Optional[str] = Query(None),
):
    require_auth(request)
    gastos = list_gastos(fuente=fuente, categorias=_parse_categorias(categorias), usuario=usuario, mes=mes, moneda=moneda)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    headers = ["Fecha", "Descripción", "Monto", "Moneda", "Fuente", "Categoría", "Usuario"]
    header_fill = PatternFill("solid", fgColor="16213E")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_i, g in enumerate(gastos, 2):
        ws.cell(row=row_i, column=1, value=g["fecha"])
        ws.cell(row=row_i, column=2, value=g["descripcion"])
        monto_cell = ws.cell(row=row_i, column=3, value=float(g["monto"]))
        monto_cell.number_format = '#,##0.00'
        ws.cell(row=row_i, column=4, value=g["moneda"])
        ws.cell(row=row_i, column=5, value=g["fuente"])
        ws.cell(row=row_i, column=6, value=g.get("categoria") or "")
        ws.cell(row=row_i, column=7, value=g.get("usuario") or "")

    col_widths = [12, 45, 14, 8, 14, 18, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"gastos_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/gastos")
def get_gastos(
    request: Request,
    fuente: Optional[str] = Query(None),
    categorias: Optional[str] = Query(None),
    usuario: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    sin_categoria: bool = Query(False),
    moneda: Optional[str] = Query(None),
    import_id: Optional[int] = Query(None),
    excluir_especiales: bool = Query(False),
):
    require_auth(request)
    return list_gastos(
        fuente=fuente,
        categorias=_parse_categorias(categorias),
        usuario=usuario,
        mes=mes,
        sin_categoria=sin_categoria,
        moneda=moneda,
        import_id=import_id,
        excluir_especiales=excluir_especiales,
    )


@router.get("/gastos/detect-transfers")
def get_detect_transfers(request: Request, days: int = Query(3)):
    require_auth(request)
    return detect_transfers(days_window=days)


@router.post("/gastos/mark-transfers")
def post_mark_transfers(body: dict, request: Request):
    require_auth(request)
    pairs = body.get("pairs", [])
    id_pairs = [(p[0], p[1]) for p in pairs if len(p) == 2]
    mark_transfers(id_pairs)
    return {"ok": True, "marcados": len(id_pairs) * 2}


@router.delete("/gastos")
def delete_all(
    request: Request,
    fuente:    Optional[str] = Query(None),
    import_id: Optional[int] = Query(None),
):
    require_auth(request)
    deleted = delete_all_gastos(fuente=fuente, import_id=import_id)
    return {"ok": True, "eliminados": deleted, "fuente": fuente, "import_id": import_id}


@router.delete("/gastos/{gasto_id}")
def delete_gasto(gasto_id: int, request: Request):
    require_auth(request)
    from fastapi import HTTPException
    if not delete_gasto_manual(gasto_id):
        raise HTTPException(403, "Solo se pueden eliminar movimientos de cuentas manuales")
    return {"ok": True}


@router.patch("/gastos/{gasto_id}/categoria")
def patch_categoria(gasto_id: int, body: dict, request: Request):
    require_auth(request)
    categoria = body.get("categoria", "")
    gasto = get_gasto(gasto_id)
    update_categoria(gasto_id, categoria)
    # Return suggestion so the frontend can prompt the user before adding to rules
    sugerencia = gasto["descripcion"] if (categoria and gasto) else None
    return {"ok": True, "sugerencia_keyword": sugerencia, "categoria": categoria}


@router.patch("/gastos/{gasto_id}/fecha")
def patch_fecha(gasto_id: int, body: dict, request: Request):
    require_auth(request)
    from fastapi import HTTPException
    fecha = (body.get("fecha") or "").strip()
    if not fecha:
        raise HTTPException(400, "fecha requerida")
    update_gasto_fecha(gasto_id, fecha)
    return {"ok": True}


@router.patch("/gastos/{gasto_id}/usuario")
def patch_usuario(gasto_id: int, body: dict, request: Request):
    require_auth(request)
    update_usuario(gasto_id, body.get("usuario", ""))
    return {"ok": True}
